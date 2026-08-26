#!/usr/bin/env python3
"""
Build the "live" Excel workbook operators open to see the transactions table.

The workbook embeds a Power Query that pulls ``GET /api/transactions/export.csv``
(auth via the ``X-Export-Key`` header) and is configured to:
  * refresh on open  → a closed workbook is fresh the moment it is opened;
  * background-refresh every N minutes → an open workbook keeps pulling new receipts.

Design constraints that shape the output (see the endpoint for the server side):
  * rows arrive ``id ASC`` (append-only) so a manual highlight never slips to
    another receipt on refresh;
  * ``preserveFormatting`` on / ``adjustColumnWidth`` off so operator fonts, fills
    and column widths survive every refresh;
  * base look rides on a Table Style (self-heals + extends to new rows);
  * the first column is the immutable ``id`` — the stable key for notes/joins.

Why hand-built OOXML: openpyxl silently drops connections/queryTables/customXml on
save, so we let it produce a valid base shell (theme, styles, sharedStrings) and then
inject the Power Query parts into the zip ourselves. The M code lives base64-encoded
inside a DataMashup blob (MS-QDEFF); permission bindings are DPAPI/machine-bound and
therefore written empty — Excel re-derives them on first save (this is normal and does
NOT trigger a repair prompt).

Usage:
    python build_live_excel.py --url https://host/api/transactions/export.csv \
        --key THE_EXPORT_KEY --out receipts_live.xlsx [--interval 1]
"""
from __future__ import annotations

import argparse
import base64
import io
import re
import struct
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Column headers, in order, exactly as GET /api/transactions/export.csv emits them:
# id first, then EXPORT_COLUMN_HEADERS in EXPORT_DEFAULT_COLUMN_KEYS order
# (backend/api/routes/transactions.py). Keep in sync with that endpoint.
HEADERS = [
    "id",
    "№",
    "№ опер.",
    "Дата и Время",
    "Дата",
    "Время",
    "День",
    "Оператор/Продавец",
    "Приложение",
    "Получатель",
    "Карта получателя",
    "Сумма",
    "Остаток",
    "ПК",
    "П2П",
    "Тип",
    "Валюта",
    "Источник",
]

# Column widths (Excel width units) keyed by header.
WIDTHS = {
    "id": 9.0,
    "№": 6.0,
    "№ опер.": 9.0,
    "Дата и Время": 20.0,
    "Дата": 12.0,
    "Время": 10.0,
    "День": 7.0,
    "Оператор/Продавец": 28.0,
    "Приложение": 18.0,
    "Получатель": 24.0,
    "Карта получателя": 18.0,
    "Сумма": 16.0,
    "Остаток": 16.0,
    "ПК": 7.0,
    "П2П": 6.0,
    "Тип": 14.0,
    "Валюта": 9.0,
    "Источник": 12.0,
}

# Headers coerced to real numbers in Power Query so operators can SUM/filter them.
NUMERIC_TEXT = ["Сумма", "Остаток"]

QUERY_NAME = "ExportData"
SHEET_TITLE = "Транзакции"
# A fixed GUID for the customXml datastore item (any stable GUID is fine).
DATASTORE_ITEM_ID = "{4A3F1C22-8E5B-4C9A-9F2D-1B7E6A0C3D45}"

CT = "http://schemas.openxmlformats.org/package/2006/content-types"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _u32(n: int) -> bytes:
    return struct.pack("<I", n)


def _m_string(value: str) -> str:
    """Escape a Python string for a Power Query M string literal."""
    return value.replace('"', '""')


def build_m_code(url: str, key: str) -> str:
    type_pairs = ", ".join(f'{{"{name}", type number}}' for name in NUMERIC_TEXT)
    return (
        "section Section1;\r\n"
        "\r\n"
        f'shared {QUERY_NAME} = let\r\n'
        f'    Source = Csv.Document(Web.Contents("{_m_string(url)}", '
        f'[Headers = [#"X-Export-Key" = "{_m_string(key)}"]]), '
        f"[Delimiter = \",\", Columns = {len(HEADERS)}, Encoding = 65001, "
        "QuoteStyle = QuoteStyle.Csv]),\r\n"
        "    Promoted = Table.PromoteHeaders(Source, [PromoteAllScalars = true]),\r\n"
        '    Typed = Table.TransformColumnTypes(Promoted, {{"id", Int64.Type}, '
        f"{type_pairs}}})\r\n"
        "in\r\n"
        "    Typed;\r\n"
    )


def build_datamashup(m_code: str) -> str:
    """Return the base64 DataMashup blob (MS-QDEFF) carrying the M query."""
    # -- packageParts: an OPC zip with the section document ------------------
    parts_buf = io.BytesIO()
    with zipfile.ZipFile(parts_buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="utf-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="xml" ContentType="text/xml" />'
            '<Default Extension="m" ContentType="text/plain" /></Types>',
        )
        z.writestr(
            "Config/Package.xml",
            '<?xml version="1.0" standalone="no"?>\r\n'
            '<Package xmlns:xsd="http://www.w3.org/2001/XMLSchema" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            "<Version>2.130.527.0</Version><MinVersion>2.21.0.0</MinVersion>"
            "<Culture>en-US</Culture></Package>",
        )
        z.writestr("Formulas/Section1.m", m_code.encode("utf-8"))
    package_parts = parts_buf.getvalue()

    # -- permissions (firewall off: single web source, skip privacy prompt) --
    permissions = (
        '﻿<?xml version="1.0" encoding="utf-8"?>'
        '<PermissionList xmlns:xsd="http://www.w3.org/2001/XMLSchema" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<CanEvaluateFuturePackages>false</CanEvaluateFuturePackages>"
        "<FirewallEnabled>false</FirewallEnabled></PermissionList>"
    ).encode("utf-8")

    # -- metadata: nested length-prefixed stream -----------------------------
    metadata_xml = (
        '﻿<?xml version="1.0" encoding="utf-8"?>'
        '<LocalPackageMetadataFile xmlns:xsd="http://www.w3.org/2001/XMLSchema" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><Items>'
        "<Item><ItemLocation><ItemType>Formula</ItemType>"
        f"<ItemPath>Section1/{QUERY_NAME}</ItemPath></ItemLocation><StableEntries>"
        '<Entry Type="FillEnabled" Value="l1" />'
        '<Entry Type="FillObjectType" Value="sTable" />'
        '<Entry Type="FillToDataModelEnabled" Value="l0" />'
        '<Entry Type="ResultType" Value="sTable" />'
        f'<Entry Type="FillTarget" Value="s{QUERY_NAME}" /></StableEntries></Item>'
        "<Item><ItemLocation><ItemType>AllFormulas</ItemType><ItemPath /></ItemLocation>"
        '<StableEntries><Entry Type="QueryGroups" Value="sAAAAAA==" />'
        '<Entry Type="Relationships" Value="sAAAAAA==" /></StableEntries></Item>'
        "</Items></LocalPackageMetadataFile>"
    ).encode("utf-8")
    empty_zip = b"PK\x05\x06" + b"\x00" * 18  # 22-byte end-of-central-directory record
    metadata = _u32(0) + _u32(len(metadata_xml)) + metadata_xml + _u32(len(empty_zip)) + empty_zip

    # -- outer stream --------------------------------------------------------
    blob = (
        _u32(0)
        + _u32(len(package_parts)) + package_parts
        + _u32(len(permissions)) + permissions
        + _u32(len(metadata)) + metadata
        + _u32(0)  # permissionBindings: empty (machine-bound DPAPI, re-derived by Excel)
    )
    return base64.b64encode(blob).decode("ascii")


def build_base_shell() -> bytes:
    """A valid, styled base workbook (header row only) via openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE
    ws.append(HEADERS)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F1F1F")
    for idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = WIDTHS.get(header, 14.0)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # A real ListObject so the base shell already carries a Table Style; the
    # injected queryTable takes it over. Ref spans header + one empty data row.
    last_col = get_column_letter(len(HEADERS))
    table = Table(displayName=QUERY_NAME, ref=f"A1:{last_col}2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _col_letter(n: int) -> str:
    return get_column_letter(n)


def _table_xml() -> str:
    last_col = _col_letter(len(HEADERS))
    columns = "".join(
        f'<tableColumn id="{i}" name="{_xml_escape(h)}" uniqueName="{i}" queryTableFieldId="{i}"/>'
        for i, h in enumerate(HEADERS, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'id="1" name="{QUERY_NAME}" displayName="{QUERY_NAME}" ref="A1:{last_col}2" '
        'headerRowCount="1" tableType="queryTable">'
        f'<autoFilter ref="A1:{last_col}2"/>'
        f'<tableColumns count="{len(HEADERS)}">{columns}</tableColumns>'
        '<tableStyleInfo name="TableStyleMedium2" showFirstColumn="0" showLastColumn="0" '
        'showRowStripes="1" showColumnStripes="0"/></table>'
    )


def _query_table_xml() -> str:
    # Minimal, Excel-clean shape (matches EPPlus output). Extra attributes here
    # (autoFormatId / apply*Formats / preserveFormatting / adjustColumnWidth) make
    # Excel reject and strip the whole part on open ("Removed Part … External data
    # range"), which kills the query. preserveFormatting defaults to true, so operator
    # fonts/fills still survive refresh without stating it.
    fields = "".join(
        f'<queryTableField id="{i}" name="{_xml_escape(h)}" tableColumnId="{i}"/>'
        for i, h in enumerate(HEADERS, start=1)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<queryTable xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'name="{QUERY_NAME}" connectionId="1" refreshOnLoad="1">'
        f'<queryTableRefresh nextId="{len(HEADERS) + 1}"><queryTableFields '
        f'count="{len(HEADERS)}">{fields}</queryTableFields></queryTableRefresh></queryTable>'
    )


def _connections_xml(interval: int) -> str:
    # Mashup connection, aligned with EPPlus's Excel-clean output. `interval` (minutes)
    # drives the periodic background refresh while the workbook is open; `refreshOnLoad`
    # on the queryTable covers refresh-on-open.
    conn_str = (
        "Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;"
        f"Location={QUERY_NAME};Extended Properties=&quot;&quot;"
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<connection id="1" name="{QUERY_NAME}" background="1" saveData="1" '
        f'refreshedVersion="8" type="5" interval="{interval}">'
        f'<dbPr connection="{conn_str}"/>'
        "</connection></connections>"
    )


def _item1_xml(datamashup_b64: str) -> str:
    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<DataMashup xmlns="http://schemas.microsoft.com/DataMashup">'
        f"{datamashup_b64}</DataMashup>"
    )


def _item_props_xml() -> str:
    return (
        '<?xml version="1.0" encoding="utf-8" standalone="no"?>'
        f'<ds:datastoreItem ds:itemID="{DATASTORE_ITEM_ID}" '
        'xmlns:ds="http://schemas.openxmlformats.org/officeDocument/2006/customXml">'
        '<ds:schemaRefs><ds:schemaRef ds:uri="http://schemas.microsoft.com/DataMashup"/>'
        "</ds:schemaRefs></ds:datastoreItem>"
    )


def _xml_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    )


def inject_power_query(base_zip: bytes, url: str, key: str, interval: int) -> bytes:
    """
    Take openpyxl's base workbook (which already wires sheet → table1 via its own
    rels) and layer the Power Query machinery on top: rewrite table1.xml as a
    queryTable-backed table, add its queryTable + connection + DataMashup parts, and
    register them in [Content_Types].xml and the workbook rels. The sheet and its rels
    are left untouched — openpyxl's tablePart r:id already points at table1.xml.
    """
    datamashup = build_datamashup(build_m_code(url, key))
    zin = zipfile.ZipFile(io.BytesIO(base_zip), "r")
    names = set(zin.namelist())

    if "xl/tables/table1.xml" not in names:
        raise RuntimeError(f"unexpected base workbook layout: {sorted(names)}")

    # --- patch [Content_Types].xml: add PQ part overrides -------------------
    content_types = zin.read("[Content_Types].xml").decode("utf-8")
    overrides = (
        '<Override PartName="/xl/connections.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"/>'
        '<Override PartName="/xl/queryTables/queryTable1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml"/>'
        '<Override PartName="/customXml/itemProps1.xml" ContentType="application/vnd.openxmlformats-officedocument.customXmlProperties+xml"/>'
    )
    if "/xl/tables/table1.xml" not in content_types:  # openpyxl usually declares this already
        overrides += '<Override PartName="/xl/tables/table1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"/>'
    content_types = content_types.replace("</Types>", overrides + "</Types>")

    # --- patch workbook.xml.rels: connections + customXml -------------------
    wb_rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    existing_ids = [int(m) for m in re.findall(r'Id="rId(\d+)"', wb_rels)]
    next_id = (max(existing_ids) + 1) if existing_ids else 1
    rid_conn, rid_customxml = f"rId{next_id}", f"rId{next_id + 1}"
    wb_rels = wb_rels.replace(
        "</Relationships>",
        f'<Relationship Id="{rid_conn}" Type="{REL}/connections" Target="connections.xml"/>'
        f'<Relationship Id="{rid_customxml}" Type="{REL}/customXml" Target="/customXml/item1.xml"/>'
        "</Relationships>",
    )

    # --- workbook.xml: hidden defined name for the external data range ------
    # Query tables require a (hidden) defined name covering the data range; without
    # it Excel strips the queryTable part on open ("Removed Part … External data
    # range"). This was the missing piece vs an EPPlus-produced reference file.
    last_col = _col_letter(len(HEADERS))
    workbook_xml = zin.read("xl/workbook.xml").decode("utf-8")
    defined_name = (
        f'<definedNames><definedName name="{QUERY_NAME}" localSheetId="0" hidden="1">'
        f"{SHEET_TITLE}!$A$1:${last_col}$2</definedName></definedNames>"
    )
    if "<definedNames />" in workbook_xml:
        workbook_xml = workbook_xml.replace("<definedNames />", defined_name, 1)
    elif "<definedNames/>" in workbook_xml:
        workbook_xml = workbook_xml.replace("<definedNames/>", defined_name, 1)
    else:
        workbook_xml = workbook_xml.replace("<calcPr", defined_name + "<calcPr", 1)

    # --- write the final package -------------------------------------------
    # Note: the sheet keeps dimension A1:<last>1 with only the header row while the
    # table ref is A1:<last>2 — Excel accepts that (EPPlus emits exactly this shape).
    out = io.BytesIO()
    patched = {
        "[Content_Types].xml": content_types,
        "xl/_rels/workbook.xml.rels": wb_rels,
        "xl/workbook.xml": workbook_xml,
        "xl/tables/table1.xml": _table_xml(),
    }
    added = {
        "xl/tables/_rels/table1.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{REL}/queryTable" Target="/xl/queryTables/queryTable1.xml"/>'
            "</Relationships>"
        ),
        "xl/queryTables/queryTable1.xml": _query_table_xml(),
        "xl/connections.xml": _connections_xml(interval),
        "customXml/item1.xml": _item1_xml(datamashup),
        "customXml/itemProps1.xml": _item_props_xml(),
        "customXml/_rels/item1.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="{REL}/customXmlProps" Target="/customXml/itemProps1.xml"/>'
            "</Relationships>"
        ),
    }
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            zout.writestr(name, patched.get(name) or zin.read(name))
        for name, data in added.items():
            zout.writestr(name, data)
    zin.close()
    return out.getvalue()


def build(url: str, key: str, interval: int) -> bytes:
    return inject_power_query(build_base_shell(), url, key, interval)


def main() -> None:
    ap = argparse.ArgumentParser(description="Build the live transactions Excel workbook.")
    ap.add_argument("--url", required=True, help="Full URL of /api/transactions/export.csv")
    ap.add_argument("--key", required=True, help="X-Export-Key value")
    ap.add_argument("--out", required=True, help="Output .xlsx path")
    ap.add_argument("--interval", type=int, default=1, help="Background refresh minutes (min 1)")
    args = ap.parse_args()
    data = build(args.url, args.key, max(1, args.interval))
    with open(args.out, "wb") as fh:
        fh.write(data)
    print(f"wrote {args.out} ({len(data)} bytes)")


if __name__ == "__main__":
    main()
