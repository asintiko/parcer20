"""
Transaction API routes
Server-side pagination, sorting, and filtering for financial transactions
"""
import os
import json
import logging
import re
from contextlib import suppress
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional

import pytz
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, asc, desc, extract, func, or_
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field

from database.connection import get_db_session
from database.models import (
    AccessScope,
    Check,
    DuplicateMergeLog,
    HiddenBotChat,
    MonitoredBotChat,
    OperatorReference,
    ReceiptProcessingTask,
    Transaction,
)
from services.telegram_tdlib_manager import TelegramTDLibManager, get_tdlib_manager
from services.receipt_processor import process_tdlib_message
from api.dependencies import (
    get_allowed_sources_for_user,
    get_effective_folder_scope_for_user,
    is_date_allowed_for_user,
    merge_scope_payloads,
    require_sources_scope,
    require_tab_access,
    require_transactions_scope,
)
from services.access_control_service import (
    clamp_requested_range,
    get_scope_window,
    is_datetime_allowed,
    write_audit_log,
)
from services.access_control_service import has_active_scopes, years_from_json
from services.root_access_config_service import get_config_scopes
from services.period_lock_service import period_lock_service
from services.auth_bot_service import publish_auth_event

# Normalization helpers
EMOJI_PATTERN = re.compile(r"[\U0001F300-\U0001FAFF\U00002600-\U000027BF]")

SOURCE_DISPLAY_MAP = {
    "TELEGRAM": "Телеграм",
    "SMS": "СМС",
    "MANUAL": "Ручной",
}
TELEGRAM_SOURCE_TYPES = ("AUTO", "TELEGRAM", "USERBOT", "BOT")

TRANSACTION_TYPE_DISPLAY_MAP = {
    "DEBIT": "Списание",
    "CREDIT": "Пополнение",
    "CONVERSION": "Конверсия",
    "REVERSAL": "Отмена",
}

logger = logging.getLogger(__name__)


def _request_ip(request: Optional[Request]) -> str:
    if request and request.client and request.client.host:
        return request.client.host
    return "unknown"


def _audit_transaction_action(
    db: Session,
    request: Optional[Request],
    current_user: Optional[Dict[str, Any]],
    *,
    action: str,
    success: bool,
    details: Optional[Dict[str, Any]] = None,
) -> None:
    try:
        write_audit_log(
            db,
            action=action,
            success=success,
            user_id=int((current_user or {}).get("id") or (current_user or {}).get("user_id") or 0),
            ip_address=_request_ip(request),
            details=details or {},
        )
    except Exception:
        logger.warning("Failed to write transaction audit: action=%s", action, exc_info=True)


async def _publish_sync_update(action: str, transaction_ids: List[int]) -> None:
    ids = [int(v) for v in transaction_ids if int(v) > 0]
    if not ids:
        return
    with suppress(Exception):
        await publish_auth_event(
            "sync_update",
            {
                "table": "transactions",
                "action": str(action),
                "ids": ids,
                "count": len(ids),
            },
        )


def normalize_source_type(source_type: Optional[str]) -> str:
    """
    Normalize source type to AUTO|MANUAL|SMS.
    """
    if not source_type:
        return "MANUAL"
    value = source_type.strip().upper()
    if value == "SMS":
        return "SMS"
    if value in {"AUTO", "TELEGRAM", "USERBOT", "BOT"}:
        return "AUTO"
    return "MANUAL"


def detect_source_channel(source_type: Optional[str], raw_text: Optional[str]) -> str:
    """
    Determine source channel for UI: TELEGRAM, SMS, or MANUAL.
    """
    if source_type:
        st = source_type.upper()
        if st == "SMS":
            return "SMS"
        if st == "MANUAL":
            return "MANUAL"
        if st in ("AUTO", "TELEGRAM", "USERBOT", "BOT"):
            return "TELEGRAM"
    if raw_text and EMOJI_PATTERN.search(raw_text):
        return "TELEGRAM"
    return "SMS"


def infer_transaction_type(raw_type: Optional[str], raw_text: Optional[str]) -> str:
    """
    Infer canonical transaction type using raw type and text with priority:
    REVERSAL > CONVERSION > explicit sign > keyword mapping > fallback DEBIT.
    """
    combined_upper = " ".join(
        part for part in [raw_type or "", raw_text or ""] if part
    ).upper()

    reversal_keywords = {"REVERSAL", "ОТМЕНА", "OTMENA", "CANCEL"}
    if any(keyword in combined_upper for keyword in reversal_keywords):
        return "REVERSAL"

    conversion_keywords = {"CONVERSION", "КОНВЕРСИЯ", "KONVERSIY", "KONVERS"}
    if any(keyword in combined_upper for keyword in conversion_keywords):
        return "CONVERSION"

    sign_text = f"{raw_type or ''} {raw_text or ''}"
    if "➕" in sign_text:
        return "CREDIT"
    if "➖" in sign_text:
        return "DEBIT"

    credit_keywords = {
        "CREDIT",
        "ПОПОЛНЕНИЕ",
        "POPOLNENIE",
        "KIRIM",
        "ПОСТУПЛЕНИЕ",
        "POSTUPLENIE",
    }
    if any(keyword in combined_upper for keyword in credit_keywords):
        return "CREDIT"

    debit_keywords = {
        "DEBIT",
        "СПИСАНИЕ",
        "SPISANIE",
        "ОПЛАТА",
        "OPLATA",
        "POKUPKA",
        "PLATEZH",
        "E-COM",
    }
    if any(keyword in combined_upper for keyword in debit_keywords):
        return "DEBIT"

    return "DEBIT"


def normalize_transaction_type(raw: Optional[str], raw_text: Optional[str] = None) -> str:
    """
    Map Russian / legacy transaction types to canonical enums using inference helper.
    """
    return infer_transaction_type(raw, raw_text)


def normalize_amount_for_response(amount: Decimal) -> str:
    """
    Present amount as positive string for UI
    """
    return f"{abs(Decimal(amount))}"


def normalize_optional_amount_for_response(amount: Optional[Decimal]) -> Optional[str]:
    """
    Present optional amount as positive string for UI
    """
    if amount is None:
        return None
    return normalize_amount_for_response(amount)


def compute_weekday_label(dt: datetime) -> str:
    weekdays = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    return weekdays[dt.weekday()]


def compute_date_display(dt: datetime) -> str:
    return dt.strftime("%d.%m.%Y")


def compute_time_display(dt: datetime) -> str:
    return dt.strftime("%H:%M")


def get_source_display(channel: str) -> str:
    return SOURCE_DISPLAY_MAP.get(channel, SOURCE_DISPLAY_MAP["SMS"])


def get_transaction_type_display(tx_type: str, raw_message: Optional[str] = None) -> str:
    """
    Возвращает человеко-читаемый текст типа транзакции.
    Для уведомлений «Счет по карте изменен» хотим видеть «Изменение счета»,
    но при этом сохраняем canonical tx_type=DEBIT (ограничение схемы).
    """
    upper = (raw_message or "").upper()
    if "СЧЕТ ПО КАРТЕ ИЗМЕН" in upper or "СЧЁТ ПО КАРТЕ ИЗМЕН" in upper:
        return "Изменение счета"
    return TRANSACTION_TYPE_DISPLAY_MAP.get(tx_type, TRANSACTION_TYPE_DISPLAY_MAP["DEBIT"])

router = APIRouter()


# Pydantic schemas
class TransactionResponse(BaseModel):
    id: int
    transaction_date: datetime
    amount: str
    currency: str
    card_last_4: Optional[str]
    operator_raw: Optional[str]
    application_mapped: Optional[str]
    transaction_type: str
    transaction_type_display: str
    balance_after: Optional[str]
    receiver_name: Optional[str] = None
    receiver_card: Optional[str] = None
    source_type: str
    source_channel: str
    source_display: str
    source_chat_id: Optional[int] = None
    source_chat_title: Optional[str] = None
    source_origin_type: Optional[str] = None
    parsing_method: Optional[str]
    parsing_confidence: Optional[float]
    is_p2p: Optional[bool] = None  # Not stored on Transaction yet
    created_at: datetime
    updated_at: Optional[datetime] = None
    raw_message: Optional[str] = None

    class Config:
        from_attributes = True


class TransactionListResponse(BaseModel):
    ok: bool = True
    total: int
    page: int
    page_size: int
    items: List[TransactionResponse]
    has_more: Optional[bool] = None
    next_cursor: Optional[Dict[str, Any]] = None
    cursor_mode: bool = False
    locked_notice: Optional[str] = None
    latest_transaction_at: Optional[str] = None
    oldest_transaction_at: Optional[str] = None
    server_generated_at: Optional[str] = None


class ProcessReceiptRequest(BaseModel):
    chat_id: int
    message_id: int
    force: bool = False


class ParsingInfo(BaseModel):
    method: Optional[str] = None
    confidence: Optional[float] = None
    notes: Optional[str] = None


class ProcessReceiptResponse(BaseModel):
    created: bool
    duplicate: bool
    transaction: TransactionResponse
    parsing: ParsingInfo


class ProcessReceiptBatchRequest(BaseModel):
    chat_id: int
    message_ids: List[int]
    force: bool = False


class ProcessReceiptBatchItem(BaseModel):
    message_id: int
    success: bool
    error: Optional[str] = None
    created: Optional[bool] = None
    duplicate: Optional[bool] = None
    transaction: Optional[TransactionResponse] = None
    parsing: Optional[ParsingInfo] = None


class ProcessReceiptBatchResponse(BaseModel):
    results: List[ProcessReceiptBatchItem]


class ProcessedStatusResponse(BaseModel):
    statuses: Dict[int, bool]


class TransactionYearsItem(BaseModel):
    year: int
    count: int


class TransactionYearsResponse(BaseModel):
    items: List[TransactionYearsItem]


class TransactionsInitResponse(BaseModel):
    security_status: Dict[str, Any]
    years: List[TransactionYearsItem]
    operators: List[str]
    applications: List[str]
    telegram_sources: List[Dict[str, Any]] = []


class DuplicateMergeEventResponse(BaseModel):
    id: int
    fingerprint: Optional[str]
    existing_transaction_id: int
    source_chat_id: Optional[int]
    source_message_id: Optional[int]
    merged: bool
    candidate_method: Optional[str]
    candidate_confidence: Optional[float]
    existing_method: Optional[str]
    existing_confidence: Optional[float]
    details: Optional[Dict[str, Any]]
    created_at: Optional[datetime]


class LocateTransactionResponse(BaseModel):
    exists: bool
    route: str
    row_id: Optional[int] = None
    column_id: Optional[str] = None
    focus_mode: str = "none"
    suggested_filters: Dict[str, Any] = Field(default_factory=dict)
    sort_hint: Dict[str, Any] = Field(default_factory=dict)
    page_hint: Optional[int] = None
    cursor_hint: Optional[Dict[str, Any]] = None


def _parse_source_chat_ids(raw_values: Optional[List[str]]) -> List[int]:
    if not raw_values:
        return []
    values: List[int] = []
    seen: set[int] = set()
    for raw in raw_values:
        if raw is None:
            continue
        for part in str(raw).split(","):
            token = part.strip()
            if not token:
                continue
            try:
                value = int(token)
            except (TypeError, ValueError):
                continue
            if value == 0 or value in seen:
                continue
            seen.add(value)
            values.append(value)
    return values


def _load_source_meta_map(
    db: Session,
    chat_ids: Optional[List[int]] = None,
) -> Dict[int, Dict[str, Any]]:
    monitored_query = db.query(
        MonitoredBotChat.chat_id,
        MonitoredBotChat.chat_title,
        MonitoredBotChat.chat_type,
    )
    if chat_ids:
        monitored_query = monitored_query.filter(MonitoredBotChat.chat_id.in_(chat_ids))

    result: Dict[int, Dict[str, Any]] = {}
    for row in monitored_query.all():
        result[int(row.chat_id)] = {
            "title": (row.chat_title or "").strip() or None,
            "chat_type": (row.chat_type or "").strip() or None,
        }

    hidden_query = db.query(
        HiddenBotChat.chat_id,
        HiddenBotChat.title_snapshot,
    )
    if chat_ids:
        hidden_query = hidden_query.filter(HiddenBotChat.chat_id.in_(chat_ids))

    for row in hidden_query.all():
        chat_id = int(row.chat_id)
        current = result.setdefault(chat_id, {})
        if not current.get("title"):
            current["title"] = (row.title_snapshot or "").strip() or None
        current.setdefault("chat_type", None)
    return result


def _resolve_transaction_source_fields(
    c: Transaction,
    source_meta_map: Optional[Dict[int, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    raw_message = getattr(c, "raw_message", None) or getattr(c, "raw_text", None)
    source_type_raw = getattr(c, "source_type", None)
    if not source_type_raw:
        added_via = getattr(c, "added_via", "") or ""
        source_type_raw = "AUTO" if added_via.lower() != "manual" else "MANUAL"

    source_channel = detect_source_channel(source_type_raw, raw_message)
    source_type_val = normalize_source_type(source_type_raw)
    source_chat_id_raw = getattr(c, "source_chat_id", None)
    try:
        source_chat_id = int(source_chat_id_raw) if source_chat_id_raw not in (None, "") else None
    except (TypeError, ValueError):
        source_chat_id = None

    source_display = get_source_display(source_channel)
    source_chat_title: Optional[str] = None
    source_origin_type: Optional[str] = None

    if source_channel == "TELEGRAM" and source_chat_id not in (None, 0):
        meta = (source_meta_map or {}).get(source_chat_id) or {}
        source_chat_title = meta.get("title") or None
        source_origin_type = meta.get("chat_type") or None
        source_display = source_chat_title or f"Telegram {source_chat_id}"

    return {
        "source_type": source_type_val,
        "source_channel": source_channel,
        "source_display": source_display,
        "source_chat_id": source_chat_id,
        "source_chat_title": source_chat_title,
        "source_origin_type": source_origin_type,
    }


def build_transaction_response(
    c: Transaction,
    source_meta_map: Optional[Dict[int, Dict[str, Any]]] = None,
) -> TransactionResponse:
    parsing_method = getattr(c, "parsing_method", None)
    parsing_confidence = getattr(c, "parsing_confidence", None)
    is_gpt_parsed = getattr(c, "is_gpt_parsed", None)
    metadata_json = getattr(c, "metadata_json", None) or getattr(c, "metadata", None) if hasattr(c, "metadata") else None
    if metadata_json:
        try:
            meta = json.loads(metadata_json)
            parsing_method = meta.get("parsing_method", parsing_method)
            parsing_confidence = meta.get("parsing_confidence", parsing_confidence)
            is_gpt_parsed = meta.get("is_gpt_parsed", is_gpt_parsed)
        except Exception:
            pass

    raw_message = getattr(c, "raw_message", None) or getattr(c, "raw_text", None)
    tx_date = getattr(c, "transaction_date", None) or getattr(c, "datetime", None)
    amount_val = getattr(c, "amount", Decimal("0"))
    card4 = getattr(c, "card_last_4", None) or getattr(c, "card_last4", None)
    operator_raw = getattr(c, "operator_raw", None) or getattr(c, "operator", None)
    app_mapped = getattr(c, "application_mapped", None) or getattr(c, "app", None)
    balance_val = getattr(c, "balance_after", None) or getattr(c, "balance", None)
    canonical_type = infer_transaction_type(
        getattr(c, "transaction_type", None),
        raw_message
    )
    source_fields = _resolve_transaction_source_fields(c, source_meta_map)

    display = get_transaction_type_display(canonical_type, raw_message)

    return TransactionResponse(
        id=c.id,
        transaction_date=tx_date,
        amount=normalize_amount_for_response(amount_val),
        currency=getattr(c, "currency", "UZS"),
        card_last_4=card4,
        operator_raw=operator_raw,
        application_mapped=app_mapped,
        transaction_type=canonical_type,
        transaction_type_display=display,
        balance_after=normalize_optional_amount_for_response(balance_val),
        receiver_name=getattr(c, "receiver_name", None),
        receiver_card=getattr(c, "receiver_card", None),
        source_type=source_fields["source_type"],
        source_channel=source_fields["source_channel"],
        source_display=source_fields["source_display"],
        source_chat_id=source_fields["source_chat_id"],
        source_chat_title=source_fields["source_chat_title"],
        source_origin_type=source_fields["source_origin_type"],
        parsing_method=parsing_method,
        parsing_confidence=parsing_confidence,
        is_p2p=getattr(c, "is_p2p", None),
        created_at=getattr(c, "created_at", None),
        updated_at=getattr(c, "updated_at", None),
        raw_message=raw_message
    )


def _apply_scope_to_query(query, scope: Optional[Dict[str, Any]]):
    if not scope:
        return query

    scope_from, scope_to, scope_years = get_scope_window(scope)
    if scope_from:
        query = query.filter(Transaction.transaction_date >= scope_from)
    if scope_to:
        query = query.filter(Transaction.transaction_date <= scope_to)
    if scope_years:
        query = query.filter(extract("year", Transaction.transaction_date).in_(scope_years))
    return query


def _effective_scope(
    db: Session,
    scope: Optional[Dict[str, Any]],
    current_user: Optional[Dict[str, Any]],
) -> Optional[Dict[str, Any]]:
    user_scope = get_effective_folder_scope_for_user(current_user, db)
    return merge_scope_payloads(scope, user_scope)


def _apply_allowed_sources_to_query(query, current_user: Optional[Dict[str, Any]]):
    """Restrict a Transaction query to the operator's allowed source chats.

    get_allowed_sources_for_user returns None for admin (no restriction) and a
    set of chat_ids for an operator. Semantics match live behavior so an operator
    with no configured sources is not locked out: an EMPTY set means "no source
    restriction configured" and applies no filter; a NON-empty set restricts to
    those chats. Runs unconditionally on every read so a missing source_chat_ids
    query param cannot widen an operator's real access.
    """
    allowed = get_allowed_sources_for_user(current_user)
    if not allowed:  # None (admin) or empty set (no restriction configured)
        return query
    return query.filter(Transaction.source_chat_id.in_(allowed))


def _ensure_source_allowed(tx, current_user: Optional[Dict[str, Any]]) -> None:
    """Raise 404 if a single transaction is outside the operator's allowed sources.

    Only enforced when allowed_sources is NON-empty (see
    _apply_allowed_sources_to_query for the empty-set rationale). 404 (not 403)
    so existence of out-of-scope rows is not disclosed.
    """
    if tx is None:
        return
    allowed = get_allowed_sources_for_user(current_user)
    if not allowed:
        return
    if tx.source_chat_id not in allowed:
        raise HTTPException(status_code=404, detail="Transaction not found")


def _user_forbidden_ranges(current_user: Optional[Dict[str, Any]]) -> List[tuple[date, date]]:
    if not current_user:
        return []
    if str(current_user.get("role") or "").lower() == "admin":
        return []
    raw = current_user.get("forbidden_periods")
    if not isinstance(raw, list):
        return []
    ranges: List[tuple[date, date]] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        try:
            start = datetime.fromisoformat(str(item.get("from"))).date()
            end = datetime.fromisoformat(str(item.get("to"))).date()
        except Exception:
            continue
        if start > end:
            start, end = end, start
        ranges.append((start, end))
    return ranges


def _apply_user_forbidden_ranges_to_query(query, ranges: List[tuple[date, date]]):
    if not ranges:
        return query
    clauses = []
    for range_start, range_end in ranges:
        clauses.append(
            and_(
                func.date(Transaction.transaction_date) >= range_start,
                func.date(Transaction.transaction_date) <= range_end,
            )
        )
    if clauses:
        query = query.filter(~or_(*clauses))
    return query


def _build_allowed_ranges(
    db: Session,
    effective_date_from: Optional[datetime],
    effective_date_to: Optional[datetime],
) -> List[tuple[date, date]]:
    if not effective_date_from and not effective_date_to:
        return []
    start_dt = effective_date_from or effective_date_to
    end_dt = effective_date_to or effective_date_from
    if start_dt is None or end_dt is None:
        return []
    start_date = start_dt.date()
    end_date = end_dt.date()
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return period_lock_service.filter_locked_range(start_date, end_date, db)


def _apply_allowed_ranges_to_query(query, ranges: List[tuple[date, date]]):
    if not ranges:
        return query
    clauses = []
    for range_start, range_end in ranges:
        clauses.append(
            and_(
                func.date(Transaction.transaction_date) >= range_start,
                func.date(Transaction.transaction_date) <= range_end,
            )
        )
    return query.filter(or_(*clauses))


def _years_from_period(period_start: Optional[datetime], period_end: Optional[datetime]) -> List[int]:
    if not period_start and not period_end:
        return []
    start = period_start or period_end
    end = period_end or period_start
    if start is None or end is None:
        return []
    if start > end:
        start, end = end, start
    years: List[int] = []
    for year in range(start.year, end.year + 1):
        years.append(year)
    return years


def _apply_locked_periods_to_query(query, db: Session):
    locks = period_lock_service.get_active_locks(db)
    if not locks:
        return query
    clauses = []
    for lock in locks:
        clauses.append(
            and_(
                func.date(Transaction.transaction_date) >= lock.date_from,
                func.date(Transaction.transaction_date) <= lock.date_to,
            )
        )
    return query.filter(~or_(*clauses))


def _assert_tx_in_scope(scope: Optional[Dict[str, Any]], tx: Transaction) -> None:
    if not scope:
        return
    if not is_datetime_allowed(scope, getattr(tx, "transaction_date", None)):
        raise HTTPException(status_code=403, detail="Transaction is outside of current scope")


def _assert_not_in_locked_period(
    db: Session,
    tx_date: Optional[datetime],
    current_user: Optional[Dict[str, Any]] = None,
) -> None:
    if not tx_date:
        return
    if str((current_user or {}).get("role") or "").lower() == "admin":
        return
    user_id = (current_user or {}).get("id")
    if period_lock_service.is_date_locked(tx_date.date(), db, user_id=user_id):
        raise HTTPException(status_code=403, detail="Transaction is in a locked period")


def _assert_user_date_allowed(current_user: Optional[Dict[str, Any]], tx_date: Optional[datetime]) -> None:
    if not tx_date:
        return
    if not is_date_allowed_for_user(current_user, tx_date):
        raise HTTPException(status_code=403, detail="Transaction date is forbidden for this user")


def _is_date_in_lock(tx_date: Optional[datetime], locks: List[Any]) -> bool:
    if not tx_date:
        return False
    tx_day = tx_date.date()
    return any(lock.date_from <= tx_day <= lock.date_to for lock in locks)


@router.get("/years", response_model=TransactionYearsResponse)
async def get_transaction_years(
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    effective_scope = _effective_scope(db, scope, current_user)
    forbidden_ranges = _user_forbidden_ranges(current_user)
    query = db.query(
        extract("year", Transaction.transaction_date).label("year"),
        func.count(Transaction.id).label("count"),
    )
    query = _apply_scope_to_query(query, effective_scope)
    query = _apply_user_forbidden_ranges_to_query(query, forbidden_ranges)
    query = _apply_locked_periods_to_query(query, db)
    rows = (
        query.group_by(extract("year", Transaction.transaction_date))
        .order_by(extract("year", Transaction.transaction_date))
        .all()
    )

    items = [
        TransactionYearsItem(year=int(row.year), count=int(row.count or 0))
        for row in rows
        if row.year is not None
    ]
    return TransactionYearsResponse(items=items)


@router.get("/init", response_model=TransactionsInitResponse)
async def get_transactions_init(
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    effective_scope = _effective_scope(db, scope, current_user)
    forbidden_ranges = _user_forbidden_ranges(current_user)
    query = db.query(
        extract("year", Transaction.transaction_date).label("year"),
        func.count(Transaction.id).label("count"),
    )
    query = _apply_scope_to_query(query, effective_scope)
    query = _apply_user_forbidden_ranges_to_query(query, forbidden_ranges)
    query = _apply_locked_periods_to_query(query, db)
    year_rows = (
        query.group_by(extract("year", Transaction.transaction_date))
        .order_by(extract("year", Transaction.transaction_date))
        .all()
    )

    years_items = [
        TransactionYearsItem(year=int(row.year), count=int(row.count or 0))
        for row in year_rows
        if row.year is not None
    ]

    available_years: List[int] = []
    db_scopes = db.query(AccessScope).filter(AccessScope.is_active.is_(True)).all()
    for row in db_scopes:
        available_years.extend(years_from_json(row.years_json))
        available_years.extend(_years_from_period(row.period_start, row.period_end))
    for scope_cfg in get_config_scopes(active_only=True):
        available_years.extend([int(y) for y in (scope_cfg.get("years") or []) if str(y).isdigit()])
        start_raw = scope_cfg.get("period_start")
        end_raw = scope_cfg.get("period_end")
        try:
            start_dt = datetime.fromisoformat(str(start_raw).replace("Z", "+00:00")) if start_raw else None
        except Exception:
            start_dt = None
        try:
            end_dt = datetime.fromisoformat(str(end_raw).replace("Z", "+00:00")) if end_raw else None
        except Exception:
            end_dt = None
        available_years.extend(_years_from_period(start_dt, end_dt))
    available_years = sorted(set(available_years))

    # Ensure all scope-available years appear in years_items (with count=0 if no transactions)
    existing_years = {item.year for item in years_items}
    for scope_year in available_years:
        if scope_year not in existing_years:
            years_items.append(TransactionYearsItem(year=scope_year, count=0))
    years_items.sort(key=lambda x: x.year)

    operators = [
        str(row[0])
        for row in db.query(OperatorReference.operator_name)
        .filter(OperatorReference.is_active.is_(True))
        .distinct()
        .all()
        if row and row[0]
    ]
    applications = [
        str(row[0])
        for row in db.query(OperatorReference.application_name)
        .filter(OperatorReference.is_active.is_(True))
        .distinct()
        .all()
        if row and row[0]
    ]

    allowed_sources = get_allowed_sources_for_user(current_user)
    source_query = db.query(
        Transaction.source_chat_id.label("chat_id"),
        func.count(Transaction.id).label("count"),
    )
    source_query = _apply_scope_to_query(source_query, effective_scope)
    source_query = _apply_user_forbidden_ranges_to_query(source_query, forbidden_ranges)
    source_query = _apply_locked_periods_to_query(source_query, db)
    source_query = source_query.filter(
        Transaction.source_chat_id.isnot(None),
        Transaction.source_chat_id != 0,
        or_(
            Transaction.source_type.in_(TELEGRAM_SOURCE_TYPES),
            Transaction.source_type.is_(None),
        ),
    )
    if allowed_sources is not None:
        if not allowed_sources:
            source_rows = []
        else:
            source_rows = (
                source_query
                .filter(Transaction.source_chat_id.in_(sorted(allowed_sources)))
                .group_by(Transaction.source_chat_id)
                .order_by(func.count(Transaction.id).desc(), Transaction.source_chat_id.asc())
                .all()
            )
    else:
        source_rows = (
            source_query
            .group_by(Transaction.source_chat_id)
            .order_by(func.count(Transaction.id).desc(), Transaction.source_chat_id.asc())
            .all()
        )

    source_ids = [int(row.chat_id) for row in source_rows if row.chat_id is not None]
    source_meta_map = _load_source_meta_map(db, source_ids)
    telegram_sources = [
        {
            "chat_id": int(row.chat_id),
            "title": (source_meta_map.get(int(row.chat_id), {}).get("title") or f"Telegram {int(row.chat_id)}"),
            "chat_type": source_meta_map.get(int(row.chat_id), {}).get("chat_type"),
            "count": int(row.count or 0),
        }
        for row in source_rows
        if row.chat_id is not None
    ]

    security_status = {
        "scopes_enabled": bool(has_active_scopes(db) or get_config_scopes(active_only=True)),
        "available_years": available_years,
        "current_scope": effective_scope,
    }
    return TransactionsInitResponse(
        security_status=security_status,
        years=years_items,
        operators=sorted(set(operators)),
        applications=sorted(set(applications)),
        telegram_sources=telegram_sources,
    )


@router.get("/duplicate-events", response_model=List[DuplicateMergeEventResponse])
async def get_duplicate_events(
    limit: int = Query(200, ge=1, le=5000),
    merged: Optional[bool] = Query(None),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    effective_scope = _effective_scope(db, scope, current_user)
    forbidden_ranges = _user_forbidden_ranges(current_user)
    query = db.query(DuplicateMergeLog)
    if merged is not None:
        query = query.filter(DuplicateMergeLog.merged.is_(merged))
    if effective_scope:
        scoped_tx_query = db.query(Transaction.id)
        scoped_tx_query = _apply_scope_to_query(scoped_tx_query, effective_scope)
        scoped_tx_query = _apply_user_forbidden_ranges_to_query(scoped_tx_query, forbidden_ranges)
        scoped_tx_query = _apply_locked_periods_to_query(scoped_tx_query, db)
        query = query.filter(
            DuplicateMergeLog.existing_transaction_id.in_(scoped_tx_query.subquery())
        )
    rows = (
        query.order_by(DuplicateMergeLog.created_at.desc(), DuplicateMergeLog.id.desc())
        .limit(limit)
        .all()
    )
    if not rows:
        return []

    result: List[DuplicateMergeEventResponse] = []
    for row in rows:
        tx_id = int(row.existing_transaction_id or 0)
        details = None
        if row.details_json:
            try:
                details = json.loads(row.details_json)
            except Exception:
                details = {"raw": row.details_json}
        result.append(
            DuplicateMergeEventResponse(
                id=int(row.id),
                fingerprint=row.fingerprint,
                existing_transaction_id=tx_id,
                source_chat_id=row.source_chat_id,
                source_message_id=row.source_message_id,
                merged=bool(row.merged),
                candidate_method=row.candidate_method,
                candidate_confidence=row.candidate_confidence,
                existing_method=row.existing_method,
                existing_confidence=row.existing_confidence,
                details=details,
                created_at=row.created_at,
            )
        )
    return result



@router.post("/process-receipt", response_model=ProcessReceiptResponse)
async def process_receipt_from_telegram(
    payload: ProcessReceiptRequest,
    db: Session = Depends(get_db_session),
    manager: TelegramTDLibManager = Depends(get_tdlib_manager),
    current_user: dict = Depends(require_tab_access("dashboard")),
    source_scope: Optional[Dict[str, Any]] = Depends(require_sources_scope),
):
    """
    Process a Telegram message (chat_id, message_id) into a transaction.
    """
    return await process_tdlib_message(
        chat_id=payload.chat_id,
        message_id=payload.message_id,
        force=payload.force,
        db=db,
        manager=manager,
    )


@router.post("/process-receipt-batch", response_model=ProcessReceiptBatchResponse)
async def process_receipt_batch(
    payload: ProcessReceiptBatchRequest,
    db: Session = Depends(get_db_session),
    manager: TelegramTDLibManager = Depends(get_tdlib_manager),
    current_user: dict = Depends(require_tab_access("dashboard")),
    source_scope: Optional[Dict[str, Any]] = Depends(require_sources_scope),
):
    results: List[ProcessReceiptBatchItem] = []
    for msg_id in payload.message_ids:
        try:
            res = await process_tdlib_message(
                chat_id=payload.chat_id,
                message_id=msg_id,
                force=payload.force,
                db=db,
                manager=manager,
            )
            results.append(
                ProcessReceiptBatchItem(
                    message_id=msg_id,
                    success=True,
                    created=res.created,
                    duplicate=res.duplicate,
                    transaction=res.transaction,
                    parsing=res.parsing,
                )
            )
        except HTTPException as exc:
            results.append(
                ProcessReceiptBatchItem(
                    message_id=msg_id,
                    success=False,
                    error=exc.detail if isinstance(exc.detail, str) else str(exc.detail),
                )
            )
        except Exception as exc:  # noqa: BLE001
            results.append(
                ProcessReceiptBatchItem(
                    message_id=msg_id,
                    success=False,
                    error=str(exc),
                )
            )
    return ProcessReceiptBatchResponse(results=results)


@router.get("/processed-status", response_model=ProcessedStatusResponse)
async def get_processed_status(
    chat_id: int = Query(...),
    message_ids: str = Query(..., description="Comma-separated list of message IDs"),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    source_scope: Optional[Dict[str, Any]] = Depends(require_sources_scope),
):
    try:
        ids = [int(x) for x in message_ids.split(",") if x.strip().isdigit()]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid message_ids parameter")
    if not ids:
        return ProcessedStatusResponse(statuses={})

    rows = (
        db.query(Transaction.source_message_id)
        .filter(
            Transaction.source_chat_id == int(chat_id),
            Transaction.source_message_id.in_([int(i) for i in ids]),
        )
        .all()
    )
    found_ids = {int(r[0]) for r in rows}
    statuses = {mid: (mid in found_ids) for mid in ids}
    return ProcessedStatusResponse(statuses=statuses)


# Update/Delete schemas
class TransactionUpdateRequest(BaseModel):
    """Schema for updating transaction fields"""
    transaction_date: Optional[datetime] = None
    operator_raw: Optional[str] = Field(None, max_length=255)
    application_mapped: Optional[str] = Field(None, max_length=100)
    amount: Optional[Decimal] = Field(None, ge=0)
    balance_after: Optional[Decimal] = None
    card_last_4: Optional[str] = Field(None, pattern=r'^\d{4}$')
    receiver_name: Optional[str] = Field(None, max_length=255)
    receiver_card: Optional[str] = Field(None, pattern=r'^\d{4,32}$')
    transaction_type: Optional[str] = Field(None, pattern=r'^(DEBIT|CREDIT|CONVERSION|REVERSAL)$')
    currency: Optional[str] = Field(None, pattern=r'^(UZS|USD)$')
    source_type: Optional[str] = Field(None, pattern=r'^(AUTO|MANUAL|SMS)$')
    parsing_method: Optional[str] = None
    parsing_confidence: Optional[float] = Field(None, ge=0.0, le=1.0)
    # Kept for backward compatibility with UI but ignored server-side
    is_p2p: Optional[bool] = None

    class Config:
        json_schema_extra = {
            "example": {
                "operator_raw": "Updated Operator Name",
                "amount": "150000.00",
                "application_mapped": "Payme",
                "transaction_type": "DEBIT"
            }
        }


class TransactionUpdateResponse(BaseModel):
    """Response after successful update"""
    ok: bool = True
    success: bool
    message: str
    transaction: TransactionResponse


class DeleteResponse(BaseModel):
    """Response after deletion"""
    ok: bool = True
    success: bool
    message: str
    deleted_id: int


class BulkDeleteRequest(BaseModel):
    """Schema for bulk delete operations"""
    ids: List[int] = Field(..., min_length=1, max_length=100)


class BulkDeleteResponse(BaseModel):
    """Response for bulk delete"""
    ok: bool = True
    success: bool
    deleted_count: int
    failed_ids: List[int] = []
    errors: List[str] = []


class BulkUpdateItem(BaseModel):
    id: int
    fields: dict


class BulkUpdateRequest(BaseModel):
    updates: List[BulkUpdateItem] = Field(..., min_items=1, max_items=500)


class BulkUpdateResponse(BaseModel):
    ok: bool = True
    success: bool
    updated_count: int
    failed_ids: List[int]
    errors: List[str]


class TransactionCreateRequest(BaseModel):
    datetime: datetime
    operator: str = Field(..., min_length=1, max_length=255)
    amount: Decimal = Field(..., ge=0)
    card_last4: str = Field(..., pattern=r'^\d{4}$')
    transaction_type: str = Field(..., pattern=r'^(DEBIT|CREDIT|CONVERSION|REVERSAL)$')
    currency: str = Field(..., pattern=r'^(UZS|USD)$')
    app: Optional[str] = Field(None, max_length=100)
    balance: Optional[Decimal] = None
    is_p2p: Optional[bool] = False
    receiver_name: Optional[str] = Field(None, max_length=255)
    receiver_card: Optional[str] = Field(None, pattern=r'^\d{4,32}$')
    raw_text: Optional[str] = None


@router.patch("/bulk-update", response_model=BulkUpdateResponse)
async def bulk_update_transactions(
    request: BulkUpdateRequest,
    http_request: Request,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Apply multiple updates in one request.
    """
    allowed_fields = {
        "transaction_date",
        "operator_raw",
        "application_mapped",
        "amount",
        "balance_after",
        "card_last_4",
        "receiver_name",
        "receiver_card",
        "transaction_type",
        "currency",
        "source_type",
        "is_p2p",
    }

    failed_ids: List[int] = []
    errors: List[str] = []
    updated_count = 0
    updated_ids: List[int] = []
    effective_scope = _effective_scope(db, scope, current_user)

    try:
        request_ids = [int(item.id) for item in request.updates]
        existing_query = db.query(Transaction).filter(Transaction.id.in_(request_ids))
        existing_query = _apply_allowed_sources_to_query(existing_query, current_user)
        existing_rows = existing_query.all()
        by_id = {int(row.id): row for row in existing_rows}

        for item in request.updates:
            c = by_id.get(int(item.id))
            if not c:
                failed_ids.append(item.id)
                errors.append(f"ID {item.id} not found")
                continue
            if not is_datetime_allowed(effective_scope, c.transaction_date):
                failed_ids.append(item.id)
                errors.append(f"ID {item.id} is outside allowed scope")
                continue
            if not is_date_allowed_for_user(current_user, c.transaction_date):
                failed_ids.append(item.id)
                errors.append(f"ID {item.id} is forbidden for current user")
                continue
            if period_lock_service.is_date_locked(c.transaction_date.date(), db):
                failed_ids.append(item.id)
                errors.append(f"ID {item.id} is in locked period")
                continue

            fields = item.fields
            # reject unknown fields
            for key in list(fields.keys()):
                if key not in allowed_fields:
                    fields.pop(key, None)

            # Normalize fields
            if "transaction_type" in fields:
                c.transaction_type = infer_transaction_type(
                    fields["transaction_type"],
                    getattr(c, "raw_message", None)
                )

            if "source_type" in fields:
                c.source_type = normalize_source_type(fields["source_type"])

            if "transaction_date" in fields:
                if not is_datetime_allowed(effective_scope, fields["transaction_date"]):
                    failed_ids.append(item.id)
                    errors.append(f"ID {item.id} date is outside allowed scope")
                    continue
                if not is_date_allowed_for_user(current_user, fields["transaction_date"]):
                    failed_ids.append(item.id)
                    errors.append(f"ID {item.id} date is forbidden for current user")
                    continue
                if period_lock_service.is_date_locked(fields["transaction_date"].date(), db):
                    failed_ids.append(item.id)
                    errors.append(f"ID {item.id} date is in locked period")
                    continue
                c.transaction_date = fields["transaction_date"]

            if "operator_raw" in fields:
                c.operator_raw = fields["operator_raw"]

            if "application_mapped" in fields:
                c.application_mapped = fields["application_mapped"]

            if "currency" in fields:
                c.currency = fields["currency"]

            if "card_last_4" in fields:
                c.card_last_4 = fields["card_last_4"]

            if "is_p2p" in fields:
                c.is_p2p = fields["is_p2p"]

            if "balance_after" in fields:
                c.balance_after = fields["balance_after"]

            if "receiver_name" in fields:
                c.receiver_name = fields["receiver_name"]

            if "receiver_card" in fields:
                c.receiver_card = fields["receiver_card"]

            if "amount" in fields:
                amt = Decimal(str(fields["amount"]))
                txn_type = infer_transaction_type(
                    c.transaction_type,
                    getattr(c, "raw_message", None)
                )
                store_amount = -abs(amt) if txn_type == "DEBIT" else abs(amt)
                c.amount = store_amount

            c.updated_at = func.now()
            updated_count += 1
            updated_ids.append(int(item.id))

        db.commit()
        if updated_ids:
            _audit_transaction_action(
                db,
                http_request,
                current_user,
                action="transactions_bulk_update",
                success=True,
                details={
                    "updated_count": int(updated_count),
                    "updated_ids": updated_ids,
                    "failed_ids": [int(v) for v in failed_ids],
                },
            )
            await _publish_sync_update("updated", updated_ids)

        return BulkUpdateResponse(
            success=len(failed_ids) == 0,
            updated_count=updated_count,
            failed_ids=failed_ids,
            errors=errors
        )

    except Exception:
        db.rollback()
        logger.exception("Bulk update failed")
        raise HTTPException(status_code=500, detail="Internal server error")


def _split_csv(values: Optional[str]) -> List[str]:
    if not values:
        return []
    return [v.strip() for v in values.split(",") if v.strip()]


def _escape_like(value: str) -> str:
    return value.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


EXPORT_DEFAULT_COLUMN_KEYS: List[str] = [
    "row_number",
    "operation_number",
    "date_time",
    "transaction_date",
    "time",
    "day",
    "operator_raw",
    "application_mapped",
    "receiver_name",
    "receiver_card",
    "amount",
    "balance_after",
    "card_last_4",
    "is_p2p",
    "transaction_type",
    "currency",
    "source_type",
]

EXPORT_COLUMN_HEADERS: Dict[str, str] = {
    "row_number": "№",
    "operation_number": "№ опер.",
    "date_time": "Дата и Время",
    "transaction_date": "Дата",
    "time": "Время",
    "day": "День",
    "operator_raw": "Оператор/Продавец",
    "application_mapped": "Приложение",
    "receiver_name": "Получатель",
    "receiver_card": "Карта получателя",
    "amount": "Сумма",
    "balance_after": "Остаток",
    "card_last_4": "ПК",
    "is_p2p": "П2П",
    "transaction_type": "Тип",
    "currency": "Валюта",
    "source_type": "Источник",
}

EXPORT_COLUMN_ALIASES: Dict[str, str] = {
    "transaction_type_display": "transaction_type",
    "source_display": "source_type",
}

EXPORT_COLUMN_WIDTHS: Dict[str, float] = {
    "row_number": 6.0,
    "operation_number": 10.0,
    "date_time": 20.0,
    "transaction_date": 13.0,
    "time": 10.0,
    "day": 8.0,
    "operator_raw": 28.0,
    "application_mapped": 18.0,
    "receiver_name": 24.0,
    "receiver_card": 18.0,
    "amount": 16.0,
    "balance_after": 16.0,
    "card_last_4": 8.0,
    "is_p2p": 7.0,
    "transaction_type": 14.0,
    "currency": 10.0,
    "source_type": 12.0,
}


def _resolve_transactions_sort_column(sort_by: str):
    sort_map: dict[str, Callable] = {
        "transaction_date": Transaction.transaction_date,
        "date_time": Transaction.transaction_date,
        "time": Transaction.transaction_date,
        "day": Transaction.transaction_date,
        "operation_number": Transaction.transaction_date,
        "amount": func.abs(Transaction.amount),
        "balance_after": func.abs(Transaction.balance_after),
        "operator_raw": Transaction.operator_raw,
        "application_mapped": Transaction.application_mapped,
        "receiver_name": Transaction.receiver_name,
        "receiver_card": Transaction.receiver_card,
        "card_last_4": Transaction.card_last_4,
        "is_p2p": Transaction.is_p2p,
        "currency": Transaction.currency,
        "transaction_type": Transaction.transaction_type,
        "source_type": Transaction.source_type,
        "created_at": Transaction.created_at,
        "parsing_confidence": Transaction.parsing_confidence,
        "updated_at": Transaction.updated_at,
    }
    return sort_map.get(sort_by, Transaction.transaction_date)


def _build_filtered_transactions_query(
    db: Session,
    *,
    current_user: Dict[str, Any],
    scope: Optional[Dict[str, Any]],
    date_from: Optional[datetime],
    date_to: Optional[datetime],
    operator: Optional[str],
    operators: Optional[str],
    app: Optional[str],
    apps: Optional[str],
    amount_min: Optional[Decimal],
    amount_max: Optional[Decimal],
    parsing_method: Optional[str],
    confidence_min: Optional[float],
    confidence_max: Optional[float],
    search: Optional[str],
    source_type: Optional[str],
    source_channel: Optional[str],
    source_chat_ids: Optional[List[str]],
    transaction_type: Optional[str],
    transaction_types: Optional[str],
    currency: Optional[str],
    card: Optional[str],
    days_of_week: Optional[str],
):
    effective_scope = _effective_scope(db, scope, current_user)
    effective_date_from, effective_date_to, disjoint = clamp_requested_range(effective_scope, date_from, date_to)
    if disjoint:
        return None, None

    locked_notice: Optional[str] = None
    allowed_ranges: List[tuple[date, date]] = []
    forbidden_ranges = _user_forbidden_ranges(current_user)
    if effective_date_from or effective_date_to:
        allowed_ranges = _build_allowed_ranges(db, effective_date_from, effective_date_to)
        if not allowed_ranges:
            return None, "Запрошенный период полностью заблокирован"
        if len(allowed_ranges) > 1:
            locked_notice = "Часть запрошенного периода заблокирована"

    query = db.query(Transaction)
    query = _apply_scope_to_query(query, effective_scope)
    query = _apply_allowed_sources_to_query(query, current_user)
    query = _apply_user_forbidden_ranges_to_query(query, forbidden_ranges)
    if allowed_ranges:
        query = _apply_allowed_ranges_to_query(query, allowed_ranges)
    else:
        query = _apply_locked_periods_to_query(query, db)

    if effective_date_from:
        query = query.filter(Transaction.transaction_date >= effective_date_from)
    if effective_date_to:
        query = query.filter(Transaction.transaction_date <= effective_date_to)
    if operator:
        escaped_operator = _escape_like(operator)
        query = query.filter(Transaction.operator_raw.ilike(f"%{escaped_operator}%", escape="\\"))
    operator_list = _split_csv(operators)
    if operator_list:
        query = query.filter(Transaction.operator_raw.in_(operator_list))
    if app:
        escaped_app = _escape_like(app)
        query = query.filter(Transaction.application_mapped.ilike(f"%{escaped_app}%", escape="\\"))
    app_list = _split_csv(apps)
    if app_list:
        query = query.filter(Transaction.application_mapped.in_(app_list))
    if amount_min is not None:
        query = query.filter(func.abs(Transaction.amount) >= amount_min)
    if amount_max is not None:
        query = query.filter(func.abs(Transaction.amount) <= amount_max)
    if parsing_method:
        query = query.filter(Transaction.parsing_method == parsing_method)
    if confidence_min is not None:
        query = query.filter(Transaction.parsing_confidence >= confidence_min)
    if confidence_max is not None:
        query = query.filter(Transaction.parsing_confidence <= confidence_max)
    if search:
        normalized_search = search.strip()
        if normalized_search:
            dialect_name = getattr(getattr(db, "bind", None), "dialect", None)
            dialect_name = getattr(dialect_name, "name", "")
            if dialect_name == "postgresql":
                query = query.filter(func.lower(Transaction.raw_message).op("%")(normalized_search.lower()))
            else:
                escaped_search = _escape_like(normalized_search)
                query = query.filter(Transaction.raw_message.ilike(f"%{escaped_search}%", escape="\\"))
    if source_type:
        if str(source_type).upper() == "AUTO":
            query = query.filter(
                or_(
                    Transaction.source_type.in_(TELEGRAM_SOURCE_TYPES),
                    and_(Transaction.source_chat_id.isnot(None), Transaction.source_chat_id != 0),
                )
            )
        else:
            query = query.filter(Transaction.source_type == source_type)
    if source_channel:
        source_channel_norm = str(source_channel or "").strip().upper()
        if source_channel_norm == "SMS":
            query = query.filter(Transaction.source_type == "SMS")
        elif source_channel_norm == "MANUAL":
            query = query.filter(Transaction.source_type == "MANUAL")
        elif source_channel_norm == "TELEGRAM":
            query = query.filter(
                or_(
                    Transaction.source_type.in_(TELEGRAM_SOURCE_TYPES),
                    and_(Transaction.source_chat_id.isnot(None), Transaction.source_chat_id != 0),
                )
            )
    source_chat_id_values = _parse_source_chat_ids(source_chat_ids)
    if source_chat_id_values:
        allowed_sources = get_allowed_sources_for_user(current_user)
        if allowed_sources is not None:
            source_chat_id_values = [chat_id for chat_id in source_chat_id_values if chat_id in allowed_sources]
        if not source_chat_id_values:
            query = query.filter(Transaction.id == -1)
        else:
            query = query.filter(Transaction.source_chat_id.in_(source_chat_id_values))
    if transaction_type:
        norm_type = normalize_transaction_type(transaction_type)
        query = query.filter(Transaction.transaction_type == norm_type)
    tx_type_list = _split_csv(transaction_types)
    if tx_type_list:
        norm_list = [normalize_transaction_type(t) for t in tx_type_list]
        query = query.filter(Transaction.transaction_type.in_(norm_list))
    if currency:
        query = query.filter(Transaction.currency == currency)
    if card:
        query = query.filter(Transaction.card_last_4 == card)
    dow_values = [int(x) for x in _split_csv(days_of_week) if x.isdigit()]
    if dow_values:
        query = query.filter(extract("dow", Transaction.transaction_date).in_(dow_values))

    return query, locked_notice


def _transactions_export_max_rows() -> int:
    raw = os.getenv("TRANSACTIONS_EXPORT_MAX_ROWS", "100000")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return 100000
    return max(1, value)


@router.get("/locate", response_model=LocateTransactionResponse)
async def locate_transaction_endpoint(
    transaction_id: Optional[int] = Query(default=None, ge=1),
    search: Optional[str] = Query(default=None),
    sort_by: str = Query("transaction_date", description="Sort field"),
    sort_dir: str = Query("asc", description="Sort direction"),
    page_size: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    from services.ai_agent.navigation_service import locate_transaction

    return locate_transaction(
        db,
        current_user=current_user,
        scope=scope,
        transaction_id=transaction_id,
        search=search,
        sort_by=sort_by,
        sort_dir=sort_dir,
        page_size=page_size,
    )


def _resolve_export_columns(columns: Optional[List[str]]) -> List[str]:
    if not columns:
        return list(EXPORT_DEFAULT_COLUMN_KEYS)

    normalized: List[str] = []
    seen: set[str] = set()
    for raw in columns:
        if raw is None:
            continue
        for part in str(raw).split(","):
            key = part.strip()
            if not key:
                continue
            key = EXPORT_COLUMN_ALIASES.get(key, key)
            if key in EXPORT_COLUMN_HEADERS and key not in seen:
                seen.add(key)
                normalized.append(key)

    return normalized or list(EXPORT_DEFAULT_COLUMN_KEYS)


def _build_export_row_values(
    index: int,
    total_rows: int,
    operation_number_desc: bool,
    tx: Transaction,
    source_meta_map: Optional[Dict[int, Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    tx_date = tx.transaction_date
    day_labels = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]
    day_value = day_labels[tx_date.weekday()] if tx_date else ""

    raw_message = getattr(tx, "raw_message", None)
    canonical_type = infer_transaction_type(tx.transaction_type, raw_message)
    source_fields = _resolve_transaction_source_fields(tx, source_meta_map)

    amount = None
    balance = None
    with suppress(Exception):
        amount = float(abs(Decimal(str(tx.amount))))
    if tx.balance_after is not None:
        with suppress(Exception):
            balance = float(abs(Decimal(str(tx.balance_after))))

    return {
        "row_number": index,
        "operation_number": max(total_rows - index + 1, 0) if operation_number_desc else index,
        "date_time": tx_date.strftime("%d.%m.%Y %H:%M:%S") if tx_date else "",
        "transaction_date": tx_date.strftime("%d.%m.%Y") if tx_date else "",
        "time": tx_date.strftime("%H:%M:%S") if tx_date else "",
        "day": day_value,
        "operator_raw": tx.operator_raw or "",
        "application_mapped": tx.application_mapped or "",
        "receiver_name": tx.receiver_name or "",
        "receiver_card": tx.receiver_card or "",
        "amount": amount,
        "balance_after": balance,
        "card_last_4": tx.card_last_4 or "",
        "is_p2p": 1 if tx.is_p2p else "",
        "transaction_type": get_transaction_type_display(canonical_type, raw_message),
        "currency": tx.currency or "UZS",
        "source_type": source_fields["source_display"],
    }


def _build_transactions_export_stream(
    transactions,
    export_columns: List[str],
    total_rows: int,
    operation_number_desc: bool,
    source_meta_map: Optional[Dict[int, Dict[str, Any]]] = None,
) -> BytesIO:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="Транзакции")
    for idx, column_id in enumerate(export_columns, start=1):
        with suppress(Exception):
            worksheet.column_dimensions[get_column_letter(idx)].width = EXPORT_COLUMN_WIDTHS.get(column_id, 14.0)
    worksheet.append([EXPORT_COLUMN_HEADERS[col] for col in export_columns])
    with suppress(Exception):
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(export_columns))}1"

    for index, tx in enumerate(transactions, start=1):
        values_map = _build_export_row_values(index, total_rows, operation_number_desc, tx, source_meta_map)
        row_cells: List[Any] = []
        for key in export_columns:
            value = values_map.get(key, "")
            cell = WriteOnlyCell(worksheet, value=value)
            if key in {"amount", "balance_after"} and isinstance(value, (int, float)):
                cell.number_format = "#,##0.00"
            row_cells.append(cell)
        worksheet.append(row_cells)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _serialize_optional_datetime(value: Optional[datetime]) -> Optional[str]:
    if not isinstance(value, datetime):
        return None
    return value.isoformat()


def _query_transaction_freshness(query) -> Dict[str, Optional[str]]:
    oldest_transaction_at, latest_transaction_at = query.with_entities(
        func.min(Transaction.transaction_date),
        func.max(Transaction.transaction_date),
    ).first() or (None, None)
    return {
        "oldest_transaction_at": _serialize_optional_datetime(oldest_transaction_at),
        "latest_transaction_at": _serialize_optional_datetime(latest_transaction_at),
    }


@router.get("/", response_model=TransactionListResponse)
async def get_transactions(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(100, ge=1, le=1000, description="Items per page"),
    sort_by: str = Query("transaction_date", description="Sort field"),
    sort_dir: str = Query("asc", description="Sort direction: asc|desc"),
    date_from: Optional[datetime] = Query(None, description="Start date"),
    date_to: Optional[datetime] = Query(None, description="End date"),
    operator: Optional[str] = Query(None, description="Operator contains"),
    operators: Optional[str] = Query(None, description="Comma-separated operators for IN filter"),
    app: Optional[str] = Query(None, description="Application contains"),
    apps: Optional[str] = Query(None, description="Comma-separated apps for IN filter"),
    amount_min: Optional[Decimal] = Query(None, ge=0),
    amount_max: Optional[Decimal] = Query(None, ge=0),
    parsing_method: Optional[str] = Query(None),
    confidence_min: Optional[float] = Query(None, ge=0.0, le=1.0),
    confidence_max: Optional[float] = Query(None, ge=0.0, le=1.0),
    search: Optional[str] = Query(None, description="Free-text search in raw_message"),
    source_type: Optional[str] = Query(None, pattern="^(AUTO|MANUAL|SMS)$"),
    source_channel: Optional[str] = Query(None, pattern="^(TELEGRAM|SMS|MANUAL)$"),
    source_chat_ids: Optional[List[str]] = Query(None, description="Telegram source chat ids (CSV or repeated)"),
    transaction_type: Optional[str] = Query(None, pattern="^(DEBIT|CREDIT|CONVERSION|REVERSAL)$"),
    transaction_types: Optional[str] = Query(None, description="Comma-separated transaction types for IN filter"),
    currency: Optional[str] = Query(None, pattern="^(UZS|USD)$"),
    card: Optional[str] = Query(None, description="Filter by last 4 digits of card"),
    days_of_week: Optional[str] = Query(None, description="Comma-separated day of week numbers 0-6"),
    cursor_updated_at: Optional[datetime] = Query(
        None,
        description="Cursor datetime for keyset pagination mode (transaction_date).",
    ),
    cursor_id: Optional[int] = Query(
        None,
        ge=0,
        description="Cursor id for keyset pagination mode.",
    ),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Get paginated list of transactions (Transactions) with server-side filters and sorting.
    """
    query, locked_notice = _build_filtered_transactions_query(
        db,
        current_user=current_user,
        scope=scope,
        date_from=date_from,
        date_to=date_to,
        operator=operator,
        operators=operators,
        app=app,
        apps=apps,
        amount_min=amount_min,
        amount_max=amount_max,
        parsing_method=parsing_method,
        confidence_min=confidence_min,
        confidence_max=confidence_max,
        search=search,
        source_type=source_type,
        source_channel=source_channel,
        source_chat_ids=source_chat_ids,
        transaction_type=transaction_type,
        transaction_types=transaction_types,
        currency=currency,
        card=card,
        days_of_week=days_of_week,
    )
    if query is None:
        generated_at = datetime.utcnow().isoformat() + "Z"
        return TransactionListResponse(
            total=0,
            page=page,
            page_size=page_size,
            items=[],
            locked_notice=locked_notice,
            server_generated_at=generated_at,
        )

    cursor_mode = bool(cursor_updated_at is not None or cursor_id is not None)
    total = 0
    has_more: Optional[bool] = None
    next_cursor: Optional[Dict[str, Any]] = None
    rows: List[Transaction] = []
    freshness = _query_transaction_freshness(query)
    server_generated_at = datetime.utcnow().isoformat() + "Z"

    if cursor_mode:
        # Keyset/cursor pagination for large datasets.
        # To keep behavior deterministic and index-friendly, cursor mode
        # is anchored to (transaction_date, id) ordering.
        if sort_dir.lower() == "asc":
            if cursor_updated_at is not None and cursor_id is not None:
                query = query.filter(
                    or_(
                        Transaction.transaction_date > cursor_updated_at,
                        and_(
                            Transaction.transaction_date == cursor_updated_at,
                            Transaction.id > int(cursor_id),
                        ),
                    )
                )
            elif cursor_updated_at is not None:
                query = query.filter(Transaction.transaction_date > cursor_updated_at)
            elif cursor_id is not None:
                query = query.filter(Transaction.id > int(cursor_id))

            page_rows = (
                query.order_by(asc(Transaction.transaction_date), asc(Transaction.id))
                .limit(page_size + 1)
                .all()
            )
        else:
            if cursor_updated_at is not None and cursor_id is not None:
                query = query.filter(
                    or_(
                        Transaction.transaction_date < cursor_updated_at,
                        and_(
                            Transaction.transaction_date == cursor_updated_at,
                            Transaction.id < int(cursor_id),
                        ),
                    )
                )
            elif cursor_updated_at is not None:
                query = query.filter(Transaction.transaction_date < cursor_updated_at)
            elif cursor_id is not None:
                query = query.filter(Transaction.id < int(cursor_id))

            page_rows = (
                query.order_by(desc(Transaction.transaction_date), desc(Transaction.id))
                .limit(page_size + 1)
                .all()
            )

        has_more = len(page_rows) > page_size
        if has_more:
            page_rows = page_rows[:page_size]
        rows = page_rows
        total = len(rows)
        if rows:
            last = rows[-1]
            next_cursor = {
                "id": int(last.id),
                "updated_at": (
                    last.transaction_date.isoformat()
                    if isinstance(last.transaction_date, datetime)
                    else None
                ),
            }
    else:
        sort_column = _resolve_transactions_sort_column(sort_by)
        order_fn = desc if sort_dir.lower() == "desc" else asc
        tie_breaker = desc(Transaction.id) if sort_dir.lower() == "desc" else asc(Transaction.id)
        query = query.order_by(order_fn(sort_column), tie_breaker)

        # Offset pagination (legacy/default).
        offset = (page - 1) * page_size
        rows_with_total = (
            query.add_columns(func.count(Transaction.id).over().label("_total_count"))
            .offset(offset)
            .limit(page_size)
            .all()
        )
        rows = [row for row, _ in rows_with_total]
        total = int(rows_with_total[0][1]) if rows_with_total else 0
        has_more = (offset + len(rows)) < total

    source_ids = sorted({
        int(row.source_chat_id)
        for row in rows
        if getattr(row, "source_chat_id", None) not in (None, 0) and normalize_source_type(getattr(row, "source_type", None)) == "AUTO"
    })
    source_meta_map = _load_source_meta_map(db, source_ids)
    items: List[TransactionResponse] = [build_transaction_response(c, source_meta_map) for c in rows]

    return TransactionListResponse(
        total=total,
        page=page,
        page_size=page_size,
        items=items,
        has_more=has_more,
        next_cursor=next_cursor,
        cursor_mode=cursor_mode,
        locked_notice=locked_notice,
        latest_transaction_at=freshness["latest_transaction_at"],
        oldest_transaction_at=freshness["oldest_transaction_at"],
        server_generated_at=server_generated_at,
    )


@router.get("/export")
async def export_transactions(
    sort_by: str = Query("transaction_date", description="Sort field"),
    sort_dir: str = Query("asc", description="Sort direction: asc|desc"),
    date_from: Optional[datetime] = Query(None, description="Start date"),
    date_to: Optional[datetime] = Query(None, description="End date"),
    operator: Optional[str] = Query(None, description="Operator contains"),
    operators: Optional[str] = Query(None, description="Comma-separated operators for IN filter"),
    app: Optional[str] = Query(None, description="Application contains"),
    apps: Optional[str] = Query(None, description="Comma-separated apps for IN filter"),
    amount_min: Optional[Decimal] = Query(None, ge=0),
    amount_max: Optional[Decimal] = Query(None, ge=0),
    parsing_method: Optional[str] = Query(None),
    confidence_min: Optional[float] = Query(None, ge=0.0, le=1.0),
    confidence_max: Optional[float] = Query(None, ge=0.0, le=1.0),
    search: Optional[str] = Query(None, description="Free-text search in raw_message"),
    source_type: Optional[str] = Query(None, pattern="^(AUTO|MANUAL|SMS)$"),
    source_channel: Optional[str] = Query(None, pattern="^(TELEGRAM|SMS|MANUAL)$"),
    source_chat_ids: Optional[List[str]] = Query(None, description="Telegram source chat ids (CSV or repeated)"),
    transaction_type: Optional[str] = Query(None, pattern="^(DEBIT|CREDIT|CONVERSION|REVERSAL)$"),
    transaction_types: Optional[str] = Query(None, description="Comma-separated transaction types for IN filter"),
    currency: Optional[str] = Query(None, pattern="^(UZS|USD)$"),
    card: Optional[str] = Query(None, description="Filter by last 4 digits of card"),
    days_of_week: Optional[str] = Query(None, description="Comma-separated day of week numbers 0-6"),
    columns: Optional[List[str]] = Query(
        None,
        description="Repeated or comma-separated export column ids in desired order",
    ),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    query, _ = _build_filtered_transactions_query(
        db,
        current_user=current_user,
        scope=scope,
        date_from=date_from,
        date_to=date_to,
        operator=operator,
        operators=operators,
        app=app,
        apps=apps,
        amount_min=amount_min,
        amount_max=amount_max,
        parsing_method=parsing_method,
        confidence_min=confidence_min,
        confidence_max=confidence_max,
        search=search,
        source_type=source_type,
        source_channel=source_channel,
        source_chat_ids=source_chat_ids,
        transaction_type=transaction_type,
        transaction_types=transaction_types,
        currency=currency,
        card=card,
        days_of_week=days_of_week,
    )

    max_rows = _transactions_export_max_rows()
    if query is None:
        total_rows = 0
        ordered_query = []
    else:
        total_rows = int(query.order_by(None).count())
        if total_rows > max_rows:
            raise HTTPException(
                status_code=413,
                detail={
                    "error": "export_limit_exceeded",
                    "limit": max_rows,
                    "total": total_rows,
                },
            )
        sort_column = _resolve_transactions_sort_column(sort_by)
        order_fn = desc if sort_dir.lower() == "desc" else asc
        tie_breaker = desc(Transaction.id) if sort_dir.lower() == "desc" else asc(Transaction.id)
        ordered_query = query.order_by(order_fn(sort_column), tie_breaker).yield_per(1000)

    source_meta_map = _load_source_meta_map(db)
    export_stream = _build_transactions_export_stream(
        ordered_query,
        _resolve_export_columns(columns),
        total_rows,
        sort_by == "transaction_date" and sort_dir.lower() == "desc",
        source_meta_map,
    )
    tz_name = os.getenv("TIMEZONE", "Asia/Tashkent")
    try:
        now_dt = datetime.now(pytz.timezone(tz_name))
    except Exception:
        now_dt = datetime.utcnow()
    filename = f"transactions_export_{now_dt.strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        export_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Total": str(total_rows),
        },
    )


@router.post("/", response_model=TransactionResponse)
async def create_transaction(
    payload: TransactionCreateRequest,
    request: Request,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Manually create a new transaction/check.
    """
    try:
        txn_type = infer_transaction_type(payload.transaction_type, payload.raw_text)
        amount = Decimal(str(payload.amount))
        store_amount = -abs(amount) if txn_type == "DEBIT" else abs(amount)

        dt = payload.datetime
        effective_scope = _effective_scope(db, scope, current_user)
        if not is_datetime_allowed(effective_scope, dt):
            raise HTTPException(status_code=403, detail="Transaction date is outside of allowed scope")
        _assert_user_date_allowed(current_user, dt)
        _assert_not_in_locked_period(db, dt)

        # Generate raw_message from payload if not provided
        raw_message = payload.raw_text or f"{txn_type}\n{payload.operator}\nСумма: {amount} {payload.currency}\nКарта: *{payload.card_last4}"

        txn = Transaction(
            transaction_date=dt,
            operator_raw=payload.operator,
            application_mapped=payload.app,
            amount=store_amount,
            balance_after=payload.balance,
            card_last_4=payload.card_last4,
            is_p2p=payload.is_p2p or False,
            receiver_name=payload.receiver_name,
            receiver_card=payload.receiver_card,
            transaction_type=txn_type,
            currency=payload.currency,
            source_type="MANUAL",
            source_chat_id=0,  # 0 for manual transactions
            raw_message=raw_message,
            parsing_method="REGEX_SMS",  # Use valid parsing method
            parsing_confidence=None,
            is_gpt_parsed=False,
        )

        db.add(txn)
        db.commit()
        db.refresh(txn)

        # Auto-map application if not provided
        if not payload.app and txn.operator_raw:
            try:
                from parsers.operator_mapper import OperatorMapper
                mapper = OperatorMapper(db)
                match = mapper.map_operator_details(txn.operator_raw)
                if match and match.get("application_name"):
                    txn.application_mapped = match["application_name"]
                    if match.get("is_p2p") is not None:
                        txn.is_p2p = match["is_p2p"]
                    db.commit()
                    db.refresh(txn)
            except Exception as e:
                logger.warning("Failed to auto-map operator: %s", e)

        _audit_transaction_action(
            db,
            request,
            current_user,
            action="transaction_create",
            success=True,
            details={"transaction_id": int(txn.id)},
        )
        await _publish_sync_update("created", [int(txn.id)])
        return build_transaction_response(txn)
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        logger.exception("Create transaction failed")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/{transaction_id}", response_model=TransactionResponse)
async def get_transaction(
    transaction_id: int,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """Get single transaction by ID"""
    effective_scope = _effective_scope(db, scope, current_user)
    tx = db.query(Transaction).filter(Transaction.id == transaction_id).first()

    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    _ensure_source_allowed(tx, current_user)
    _assert_tx_in_scope(effective_scope, tx)
    _assert_user_date_allowed(current_user, tx.transaction_date)
    _assert_not_in_locked_period(db, tx.transaction_date)

    source_meta_map = _load_source_meta_map(db, [int(tx.source_chat_id)] if getattr(tx, "source_chat_id", None) else None)
    return build_transaction_response(tx, source_meta_map)


@router.put("/{transaction_id}", response_model=TransactionUpdateResponse)
async def update_transaction(
    transaction_id: int,
    update_data: TransactionUpdateRequest,
    request: Request,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Update a transaction by ID (partial updates supported).
    """
    try:
        effective_scope = _effective_scope(db, scope, current_user)
        c = db.query(Transaction).filter(Transaction.id == transaction_id).first()

        if not c:
            raise HTTPException(status_code=404, detail=f"Transaction {transaction_id} not found")
        _ensure_source_allowed(c, current_user)
        _assert_tx_in_scope(effective_scope, c)
        _assert_user_date_allowed(current_user, c.transaction_date)
        _assert_not_in_locked_period(db, c.transaction_date)

        update_dict = update_data.model_dump(exclude_unset=True)

        # Normalize transaction type
        if "transaction_type" in update_dict:
            c.transaction_type = infer_transaction_type(
                update_dict["transaction_type"],
                getattr(c, "raw_message", None)
            )

        if "source_type" in update_dict and update_dict["source_type"]:
            c.source_type = normalize_source_type(update_dict["source_type"])

        if "transaction_date" in update_dict:
            if not is_datetime_allowed(effective_scope, update_dict["transaction_date"]):
                raise HTTPException(status_code=403, detail="Transaction date is outside of allowed scope")
            _assert_user_date_allowed(current_user, update_dict["transaction_date"])
            _assert_not_in_locked_period(db, update_dict["transaction_date"])
            c.transaction_date = update_dict["transaction_date"]

        if "operator_raw" in update_dict:
            c.operator_raw = update_dict["operator_raw"]

        if "application_mapped" in update_dict:
            c.application_mapped = update_dict["application_mapped"]

        if "currency" in update_dict:
            c.currency = update_dict["currency"]

        if "card_last_4" in update_dict:
            c.card_last_4 = update_dict["card_last_4"]

        if "is_p2p" in update_dict:
            c.is_p2p = update_dict["is_p2p"]

        if "balance_after" in update_dict:
            c.balance_after = update_dict["balance_after"]

        if "receiver_name" in update_dict:
            c.receiver_name = update_dict["receiver_name"]

        if "receiver_card" in update_dict:
            c.receiver_card = update_dict["receiver_card"]

        # Amount normalization: store debits as negative, credits/etc as positive
        if "amount" in update_dict:
            amt = Decimal(str(update_dict["amount"]))
            txn_type = infer_transaction_type(
                c.transaction_type,
                getattr(c, "raw_message", None)
            )
            store_amount = -abs(amt) if txn_type == "DEBIT" else abs(amt)
            c.amount = store_amount

        c.updated_at = func.now()

        db.commit()
        db.refresh(c)
        _audit_transaction_action(
            db,
            request,
            current_user,
            action="transaction_update",
            success=True,
            details={
                "transaction_id": int(c.id),
                "updated_fields": sorted(update_dict.keys()),
            },
        )
        await _publish_sync_update("updated", [int(c.id)])

        return TransactionUpdateResponse(
            success=True,
            message="Transaction updated successfully",
            transaction=build_transaction_response(
                c,
                _load_source_meta_map(db, [int(c.source_chat_id)] if getattr(c, "source_chat_id", None) else None),
            )
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        logger.exception("Update transaction failed")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.delete("/{transaction_id}", response_model=DeleteResponse)
async def delete_transaction(
    transaction_id: int,
    request: Request,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Delete a transaction by ID
    """
    try:
        effective_scope = _effective_scope(db, scope, current_user)
        transaction = db.query(Transaction).filter(Transaction.id == transaction_id).first()

        if not transaction:
            raise HTTPException(status_code=404, detail=f"Transaction {transaction_id} not found")
        _ensure_source_allowed(transaction, current_user)
        _assert_tx_in_scope(effective_scope, transaction)
        _assert_user_date_allowed(current_user, transaction.transaction_date)
        _assert_not_in_locked_period(db, transaction.transaction_date, current_user)

        # Delete related tracking records to keep sync with Telegram client
        db.query(ReceiptProcessingTask).filter(
            ReceiptProcessingTask.transaction_id == transaction_id
        ).delete(synchronize_session=False)

        db.delete(transaction)
        db.commit()
        _audit_transaction_action(
            db,
            request,
            current_user,
            action="transaction_delete",
            success=True,
            details={"transaction_id": int(transaction_id)},
        )
        await _publish_sync_update("deleted", [int(transaction_id)])

        return DeleteResponse(
            success=True,
            message="Transaction deleted successfully",
            deleted_id=transaction_id
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        logger.exception("Delete transaction failed")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/bulk-delete", response_model=BulkDeleteResponse)
async def bulk_delete_transactions(
    request: BulkDeleteRequest,
    http_request: Request,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(require_tab_access("dashboard")),
    scope: Optional[Dict[str, Any]] = Depends(require_transactions_scope),
):
    """
    Delete multiple transactions at once
    """
    try:
        effective_scope = _effective_scope(db, scope, current_user)
        ids = request.ids
        if len(ids) > 500:
            raise HTTPException(status_code=400, detail="Maximum 500 transaction IDs per request")
        rows_query = db.query(Transaction).filter(Transaction.id.in_(ids))
        rows_query = _apply_allowed_sources_to_query(rows_query, current_user)
        rows = rows_query.all()
        existing_map = {int(tx.id): tx for tx in rows}
        failed_ids = [i for i in ids if i not in existing_map]
        errors: List[str] = [f"ID {fid} not found" for fid in failed_ids]

        user_id = current_user.get("id")
        is_admin = str(current_user.get("role") or "").lower() == "admin"

        allowed_ids: List[int] = []
        for tx_id, tx in existing_map.items():
            if not is_datetime_allowed(effective_scope, tx.transaction_date):
                failed_ids.append(tx_id)
                errors.append(f"ID {tx_id} is outside current scope")
                continue
            if not is_date_allowed_for_user(current_user, tx.transaction_date):
                failed_ids.append(tx_id)
                errors.append(f"ID {tx_id} date is forbidden for current user")
                continue
            if (
                not is_admin
                and tx.transaction_date
                and period_lock_service.is_date_locked(tx.transaction_date.date(), db, user_id=user_id)
            ):
                failed_ids.append(tx_id)
                errors.append(f"ID {tx_id} is in locked period")
                continue
            allowed_ids.append(tx_id)

        deleted_count = 0
        if allowed_ids:
            allowed_set = set(allowed_ids)
            db.query(ReceiptProcessingTask).filter(
                ReceiptProcessingTask.transaction_id.in_(allowed_ids)
            ).delete(synchronize_session=False)
            for tx_id, tx in existing_map.items():
                if tx_id in allowed_set:
                    db.delete(tx)
                    deleted_count += 1
        db.commit()
        _audit_transaction_action(
            db,
            http_request,
            current_user,
            action="transactions_bulk_delete",
            success=True,
            details={
                "requested_count": len(ids),
                "deleted_count": int(deleted_count),
                "deleted_ids": [int(v) for v in allowed_ids],
                "failed_ids": [int(v) for v in failed_ids],
            },
        )
        if allowed_ids:
            await _publish_sync_update("deleted", [int(v) for v in allowed_ids])

        return BulkDeleteResponse(
            success=len(failed_ids) == 0,
            deleted_count=deleted_count,
            failed_ids=failed_ids,
            errors=errors,
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        logger.exception("Bulk delete failed")
        raise HTTPException(status_code=500, detail="Internal server error")
