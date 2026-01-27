"""
Operator Reference API routes
CRUD operations for operator/seller reference dictionary
"""
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc, or_
from typing import List, Optional
from pydantic import BaseModel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

from database.connection import get_db_session
from database.models import OperatorReference
from api.dependencies import get_current_user

router = APIRouter()


# Pydantic schemas
class OperatorReferenceResponse(BaseModel):
    id: int
    operator_name: str
    application_name: str
    is_p2p: bool
    is_active: bool

    class Config:
        from_attributes = True


class OperatorReferenceCreate(BaseModel):
    operator_name: str
    application_name: str
    is_p2p: bool = False
    is_active: bool = True


class OperatorReferenceUpdate(BaseModel):
    operator_name: Optional[str] = None
    application_name: Optional[str] = None
    is_p2p: Optional[bool] = None
    is_active: Optional[bool] = None


class OperatorReferenceListResponse(BaseModel):
    total: int
    page: int
    page_size: int
    items: List[OperatorReferenceResponse]


@router.get("/", response_model=OperatorReferenceListResponse)
async def get_operators(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: Optional[int] = Query(50, ge=1, description="Items per page (ignored when all=true)"),
    search: Optional[str] = Query(None, description="Search in operator or app name"),
    application: Optional[str] = Query(None, description="Filter by application"),
    is_p2p: Optional[bool] = Query(None, description="Filter by P2P status"),
    is_active: Optional[bool] = Query(None, description="Filter by active status (None = all)"),
    all_param: Optional[str] = Query(None, alias="all", description="Return full list without pagination"),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Get list of operators with optional pagination"""
    # Convert string parameter to boolean (handles "true", "1", "yes", "on")
    return_all = all_param is not None and all_param.lower() in ("true", "1", "yes", "on")

    query = db.query(OperatorReference)

    # Apply filters
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            or_(
                OperatorReference.operator_name.ilike(search_pattern),
                OperatorReference.application_name.ilike(search_pattern)
            )
        )

    if application:
        query = query.filter(OperatorReference.application_name == application)

    if is_p2p is not None:
        query = query.filter(OperatorReference.is_p2p == is_p2p)

    if is_active is not None:
        query = query.filter(OperatorReference.is_active == is_active)

    # Get total count
    total = query.count()

    # Return everything when explicitly requested
    if return_all:
        items = query.order_by(desc(OperatorReference.id)).all()
        return OperatorReferenceListResponse(
            total=total,
            page=1,
            page_size=total if total > 0 else 0,
            items=items
        )

    # Apply pagination
    effective_page_size = page_size or 50
    offset = (page - 1) * effective_page_size
    items = query.order_by(desc(OperatorReference.id)).offset(offset).limit(effective_page_size).all()

    return OperatorReferenceListResponse(
        total=total,
        page=page,
        page_size=effective_page_size,
        items=items
    )


@router.post("/", response_model=OperatorReferenceResponse)
async def create_operator(
    operator: OperatorReferenceCreate,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Create new operator reference"""
    try:
        # Check for duplicates
        existing = db.query(OperatorReference).filter(
            OperatorReference.operator_name == operator.operator_name,
            OperatorReference.application_name == operator.application_name
        ).first()

        if existing:
            raise HTTPException(status_code=400, detail="Operator already exists")

        new_operator = OperatorReference(**operator.dict())
        db.add(new_operator)
        db.commit()
        db.refresh(new_operator)

        return new_operator
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error creating operator: {e}")
        raise HTTPException(status_code=500, detail="Failed to create operator")


@router.put("/{operator_id}", response_model=OperatorReferenceResponse)
async def update_operator(
    operator_id: int,
    operator: OperatorReferenceUpdate,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Update operator reference"""
    try:
        db_operator = db.query(OperatorReference).filter(OperatorReference.id == operator_id).first()

        if not db_operator:
            raise HTTPException(status_code=404, detail="Operator not found")

        # Update fields
        update_data = operator.dict(exclude_unset=True)
        for key, value in update_data.items():
            setattr(db_operator, key, value)

        db.commit()
        db.refresh(db_operator)

        return db_operator
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating operator: {e}")
        raise HTTPException(status_code=500, detail="Failed to update operator")


@router.delete("/{operator_id}")
async def delete_operator(
    operator_id: int,
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Delete operator reference"""
    try:
        db_operator = db.query(OperatorReference).filter(OperatorReference.id == operator_id).first()

        if not db_operator:
            raise HTTPException(status_code=404, detail="Operator not found")

        db.delete(db_operator)
        db.commit()

        return {"message": "Operator deleted successfully"}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error deleting operator: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete operator")


@router.get("/export/excel")
async def export_to_excel(
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Export operators to Excel file"""
    # Get all active operators
    operators = db.query(OperatorReference).filter(
        OperatorReference.is_active == True
    ).order_by(OperatorReference.application_name, OperatorReference.operator_name).all()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Операторы"

    # Headers
    headers = ["Оператор/Продавец", "Приложение", "P2P"]
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Add data
    for op in operators:
        ws.append([
            op.operator_name,
            op.application_name,
            1 if op.is_p2p else 0
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=operators.xlsx"}
    )


@router.post("/import/excel")
async def import_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Import operators from Excel file"""
    try:
        def parse_bool(value) -> bool:
            """Parse truthy values from Excel cell."""
            if value is None:
                return False
            if isinstance(value, bool):
                return value
            try:
                # Numeric truthy like 1 / 0
                if isinstance(value, (int, float)):
                    return bool(int(value))
            except (TypeError, ValueError):
                pass

            str_val = str(value).strip().lower()
            if str_val in {"1", "true", "t", "yes", "y", "да", "p2p"}:
                return True
            if str_val in {"0", "false", "f", "no", "n", ""}:
                return False
            return False

        # Read Excel file
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        imported = 0
        skipped = 0
        errors = []

        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            operator_name = str(row[0]).strip() if row[0] else None
            application_name = str(row[1]).strip() if row[1] else None
            is_p2p_raw = row[2] if len(row) > 2 else None
            is_p2p = parse_bool(is_p2p_raw)

            if not operator_name or not application_name:
                errors.append(f"Row {row_idx}: Missing operator or application name")
                skipped += 1
                continue

            # Check if exists
            existing = db.query(OperatorReference).filter(
                OperatorReference.operator_name == operator_name,
                OperatorReference.application_name == application_name
            ).first()

            if existing:
                skipped += 1
                continue

            # Create new record
            new_operator = OperatorReference(
                operator_name=operator_name,
                application_name=application_name,
                # Default to False unless explicitly marked
                is_p2p=is_p2p,
                is_active=True
            )
            db.add(new_operator)
            imported += 1

        db.commit()

        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors
        }

    except Exception as e:
        db.rollback()
        print(f"Import failed: {e}")
        raise HTTPException(status_code=400, detail="Import failed")


@router.get("/applications")
async def get_applications(
    db: Session = Depends(get_db_session),
    current_user: dict = Depends(get_current_user),
):
    """Get list of unique application names"""
    apps = db.query(OperatorReference.application_name).distinct().order_by(OperatorReference.application_name).all()
    return [app[0] for app in apps]
